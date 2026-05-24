import openpyxl
from collections import Counter

opt_wb = openpyxl.load_workbook('/workspace/user_input_files/期权实盘记录.xlsx', data_only=True)
opt_ws = opt_wb.active
opt_rows = [row for row in opt_ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in row)]
opt_headers = [str(c) if c else '' for c in opt_rows[0]]
opt_data = [[str(c) if c is not None else '' for c in row] for row in opt_rows[1:]]

stk_wb = openpyxl.load_workbook('/workspace/user_input_files/正股实盘记录.xlsx', data_only=True)
stk_ws = stk_wb.active
stk_rows = [row for row in stk_ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in row)]
stk_headers = [str(c) if c else '' for c in stk_rows[0]]
stk_data = [[str(c) if c is not None else '' for c in row] for row in stk_rows[1:]]

opt_ticker = opt_headers.index('股票')
opt_dir = opt_headers.index('买/卖')
opt_strategy = opt_headers.index('策略类型')
opt_entry = opt_headers.index('进场依据')
opt_exit = opt_headers.index('出场依据')
opt_pnl = opt_headers.index('收益值')
opt_ret = opt_headers.index('期权收益率')
opt_win = opt_headers.index('胜负')
opt_type = opt_headers.index('类型')
opt_strike = opt_headers.index('行权价')
opt_multi = opt_headers.index('组合策略')
opt_entry_date = opt_headers.index('入场日期')
opt_exit_date = opt_headers.index('平仓日期')

stk_ticker = stk_headers.index('股票')
stk_dir = stk_headers.index('买/卖')
stk_strategy = stk_headers.index('策略类型')
stk_entry = stk_headers.index('进场依据')
stk_exit = stk_headers.index('出场依据')
stk_pnl = stk_headers.index('收益值')
stk_ret = stk_headers.index('收益率')
stk_win = stk_headers.index('胜负')
stk_market = stk_headers.index('市场')
stk_entry_date = stk_headers.index('入场日期')
stk_exit_date = stk_headers.index('平仓日期')
stk_days = stk_headers.index('交易时长')

def parse_ret(s):
    s = str(s).replace('%','').strip()
    try: return float(s)
    except: return None

def parse_pnl(s):
    try: return float(str(s).replace(',','').replace('$','').replace('¥','').strip())
    except: return 0.0

# 期权分析
print("=" * 70)
print("  期权实盘记录分析")
print("=" * 70)
print(f"总记录: {len(opt_data)}条")

strategies_opt = Counter(r[opt_strategy] for r in opt_data)
print("\n【策略类型分布】")
for s, cnt in strategies_opt.most_common():
    wins = sum(1 for r in opt_data if r[opt_strategy]==s and r[opt_win]=='胜')
    total_pnl = sum(parse_pnl(r[opt_pnl]) for r in opt_data if r[opt_strategy]==s)
    rets = [parse_ret(r[opt_ret]) for r in opt_data if r[opt_strategy]==s and parse_ret(r[opt_ret]) is not None]
    avg_ret = sum(rets)/len(rets) if rets else 0
    print(f"  {s:15s} {cnt}笔 胜率{wins/cnt*100:.0f}% 均收益{avg_ret:.1f}% 总${total_pnl:,.0f}")

total_opt_pnl = sum(parse_pnl(r[opt_pnl]) for r in opt_data)
wins_opt = sum(1 for r in opt_data if r[opt_win]=='胜')
print(f"\n总收益: ${total_opt_pnl:,.0f}  胜{wins_opt} 负{len(opt_data)-wins_opt}  胜率{wins_opt/len(opt_data)*100:.0f}%")

print("\n【按策略逐条明细】")
for strategy in strategies_opt:
    rows = [r for r in opt_data if r[opt_strategy]==strategy]
    print(f"\n  ▶ {strategy} ({len(rows)}笔)")
    for r in rows:
        pnl = parse_pnl(r[opt_pnl])
        ret = parse_ret(r[opt_ret])
        ret_s = f"{ret:.2f}%" if ret is not None else "N/A"
        print(f"    {r[opt_ticker]:6s} {r[opt_dir]} {r[opt_type]} 行权{r[opt_strike]} {r[opt_multi]:4s} "
              f"建仓{r[opt_entry_date]}→平仓{r[opt_exit_date]} 收益${pnl:,.0f}({ret_s}) {r[opt_win]}")

# 正股分析
print("\n" + "=" * 70)
print("  正股实盘记录分析")
print("=" * 70)
print(f"总记录: {len(stk_data)}条")

strategies_stk = Counter(r[stk_strategy] for r in stk_data)
print("\n【策略类型分布】")
for s, cnt in strategies_stk.most_common():
    wins = sum(1 for r in stk_data if r[stk_strategy]==s and r[stk_win]=='胜')
    total_pnl = sum(parse_pnl(r[stk_pnl]) for r in stk_data if r[stk_strategy]==s)
    rets = [parse_ret(r[stk_ret]) for r in stk_data if r[stk_strategy]==s and parse_ret(r[stk_ret]) is not None]
    avg_ret = sum(rets)/len(rets) if rets else 0
    print(f"  {s:15s} {cnt}笔 胜率{wins/cnt*100:.0f}% 均收益{avg_ret:.1f}% 总${total_pnl:,.0f}")

total_stk_pnl = sum(parse_pnl(r[stk_pnl]) for r in stk_data)
wins_stk = sum(1 for r in stk_data if r[stk_win]=='胜')
print(f"\n总收益: ${total_stk_pnl:,.0f}  胜{wins_stk} 负{len(stk_data)-wins_stk}  胜率{wins_stk/len(stk_data)*100:.0f}%")

print("\n【按策略逐条明细】")
for strategy in strategies_stk:
    rows = [r for r in stk_data if r[stk_strategy]==strategy]
    print(f"\n  ▶ {strategy} ({len(rows)}笔)")
    for r in rows:
        pnl = parse_pnl(r[stk_pnl])
        ret = parse_ret(r[stk_ret])
        ret_s = f"{ret:.2f}%" if ret is not None else "N/A"
        print(f"    {r[stk_ticker]:8s} {r[stk_dir]} {r[stk_market]} "
              f"建仓{r[stk_entry_date]}→平仓{r[stk_exit_date]} 持仓{r[stk_days]}天 收益${pnl:,.0f}({ret_s}) {r[stk_win]}")